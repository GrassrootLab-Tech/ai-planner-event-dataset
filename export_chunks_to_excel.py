#!/usr/bin/env python3
"""Export event_scraped_chunks for sample_website URLs to a formatted Excel workbook.

Usage:
    pip install openpyxl
    python export_chunks_to_excel.py
    python export_chunks_to_excel.py --output output/sample_chunks.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import Settings
from db.event_scraped_chunks_repo import EventScrapedChunksRepository
from db.mongo import Mongo
from sample_website import PAGE_URLS
from tags.order import METADATA_TAG_ORDER

BASE_COLUMNS: tuple[str, ...] = (
    "page_url",
    "chunk",
    "parent_section_heading",
    "scraped_at",
    "is_usable_value",
)

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
STRIPE_FILL = PatternFill("solid", fgColor="F7F9FC")
DEFAULT_OUTPUT = Path("output/sample_chunks.xlsx")
MAX_SHEET_NAME_LEN = 31
INVALID_SHEET_CHARS = re.compile(r"[\[\]\:\*\?\/\\]")


def sanitize_excel_value(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def website_name_from_url(page_url: str) -> str:
    """Return a plain site label, e.g. catchmyparty.com -> Catchmyparty."""
    host = (urlparse(page_url).hostname or "").lower().removeprefix("www.")
    if not host:
        return "Unknown"

    if host.endswith(".co.uk"):
        label = host[: -len(".co.uk")].split(".")[-1]
    else:
        parts = host.split(".")
        label = parts[0] if len(parts) == 2 else parts[-2]

    return label.title()


def make_sheet_name(page_url: str, used_names: set[str]) -> str:
    base = INVALID_SHEET_CHARS.sub("", website_name_from_url(page_url))
    base = base[:MAX_SHEET_NAME_LEN] or "Sheet"
    candidate = base
    suffix = 2
    while candidate in used_names:
        tail = str(suffix)
        candidate = f"{base[: MAX_SHEET_NAME_LEN - len(tail)]}{tail}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def format_tag_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, list):
        return ", ".join(format_tag_value(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def format_datetime(value: datetime | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def chunk_to_row(doc: dict[str, Any]) -> dict[str, Any]:
    is_usable = doc.get("is_usable") or {}
    row = {
        "page_url": doc.get("page_url", ""),
        "chunk": doc.get("chunk", ""),
        "parent_section_heading": doc.get("parent_section_heading") or "",
        "scraped_at": format_datetime(doc.get("scraped_at")),
        "is_usable_value": is_usable.get("value", ""),
    }

    metadata_tags = doc.get("metadata_tags") or {}
    for tag_name in METADATA_TAG_ORDER:
        row[tag_name] = format_tag_value(metadata_tags.get(tag_name))

    return row


def autosize_columns(ws: Worksheet, columns: list[str]) -> None:
    widths = {
        "page_url": 42,
        "chunk": 72,
        "parent_section_heading": 28,
        "scraped_at": 20,
        "is_usable_value": 14,
    }

    for index, column_name in enumerate(columns, start=1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = widths.get(column_name, 20)


def style_sheet(ws: Worksheet, columns: list[str], row_count: int) -> None:
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    for col_idx, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx in range(2, row_count + 1):
        fill = STRIPE_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = BODY_ALIGNMENT
            cell.border = THIN_BORDER
            if fill is not None:
                cell.fill = fill


def write_sheet(ws: Worksheet, rows: list[dict[str, Any]]) -> None:
    columns = list(BASE_COLUMNS) + list(METADATA_TAG_ORDER)
    ws.append(columns)

    for row in rows:
        ws.append([sanitize_excel_value(row.get(column, "")) for column in columns])

    style_sheet(ws, columns, row_count=max(len(rows), 1) + 1)
    autosize_columns(ws, columns)


async def fetch_chunks_by_url(
    repo: EventScrapedChunksRepository,
    page_url: str,
) -> list[dict[str, Any]]:
    chunks = await repo.list_by_page_url(page_url)
    return [chunk_to_row(chunk.model_dump()) for _, chunk in chunks]


async def build_workbook(output_path: Path) -> tuple[int, int]:
    settings = Settings()
    mongo = Mongo(settings.mongo_uri, settings.mongo_db_name)
    await mongo.connect()

    try:
        repo = EventScrapedChunksRepository(
            mongo.db[settings.event_scraped_chunks_collection]
        )
        workbook = Workbook()
        workbook.remove(workbook.active)

        used_sheet_names: set[str] = set()
        total_chunks = 0

        pages = [
            str(entry["url"]).strip()
            for entry in PAGE_URLS
            if entry.get("url") and str(entry["url"]).strip()
        ]

        for page_url in pages:
            rows = await fetch_chunks_by_url(repo, page_url)
            sheet_name = make_sheet_name(page_url, used_sheet_names)
            ws = workbook.create_sheet(title=sheet_name)
            write_sheet(ws, rows)
            total_chunks += len(rows)
            print(f"{sheet_name}: {len(rows)} chunks")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return len(pages), total_chunks
    finally:
        await mongo.disconnect()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export sample_website chunks from MongoDB to Excel."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sheet_count, chunk_count = asyncio.run(build_workbook(args.output))
    print(f"Wrote {chunk_count} chunks across {sheet_count} sheets to {args.output}")


if __name__ == "__main__":
    main()
