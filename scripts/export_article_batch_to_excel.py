"""Export an article URL JSON batch to a readable Excel workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "input_urls" / "article_batches" / "batch_01.json"
DEFAULT_OUTPUT = (
    ROOT / "input_urls" / "article_batch_sheet" / "batch_01.xlsx"
)

HEADERS = ("URL", "Page Title", "Description")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
EVEN_ROW_FILL = PatternFill("solid", fgColor="EAF2F8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export an article batch JSON file to formatted Excel."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def export_batch(items: list[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Article Batch 01"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26

    for row_number, item in enumerate(items, start=2):
        sheet.append(
            [
                item.get("url", ""),
                item.get("page_title", ""),
                item.get("description", ""),
            ]
        )

        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = EVEN_ROW_FILL

        url_cell = sheet.cell(row=row_number, column=1)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"

        for cell in sheet[row_number]:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
        sheet.row_dimensions[row_number].height = 48

    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    sheet.column_dimensions[get_column_letter(1)].width = 65
    sheet.column_dimensions[get_column_letter(2)].width = 45
    sheet.column_dimensions[get_column_letter(3)].width = 90

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    items = json.loads(args.input.read_text(encoding="utf-8"))
    if not isinstance(items, list):
        raise SystemExit(f"Expected a JSON array in {args.input}")

    export_batch(items, args.output)
    print(f"Exported {len(items)} articles to {args.output}")


if __name__ == "__main__":
    main()
